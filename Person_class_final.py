from datetime import datetime, date
import time
import openpyxl


class Person:
    PEOPLE_BASE = []
    PEOPLE_AMOUNT = 0

    def __init__(self, first_name, birth_date, gender, last_name='', middle_name='', death_date=None):
        self.first_name = first_name
        self.last_name = last_name
        self.middle_name = middle_name
        self.birth_date = birth_date
        self.death_date = death_date
        self.gender = gender
        Person.PEOPLE_BASE.append(self)
        Person.PEOPLE_AMOUNT += 1
        self.number = Person.PEOPLE_AMOUNT
        self.age = Person.calc_age(self)

    def __str__(self):
        error = '*** available genders: m/male or f/female only ***'
        he_she = "he" if self.gender in ['m', 'male'] else "she" if self.gender in ['f', 'female'] else ''
        men_women = "men" if self.gender in ['m', 'male'] else "women" if self.gender in ['f', 'female'] else error
        output = f"{self.first_name.title()}"
        birth_date_formatted = format_date(self.birth_date)
        death_date_formatted = format_date(self.death_date) if self.death_date else None
        if self.last_name:
            output += f" {self.last_name.title()}"
        if self.middle_name:
            output += f" {self.middle_name.title()}"
        if self.birth_date:
            output += f' {self.calc_age()} years old, {men_women}, {he_she} was born in {birth_date_formatted}'
            if self.death_date:
                output += f',  {he_she} died in {death_date_formatted}'
        return output

    def calc_age(self):
        try:
            if self.birth_date:
                birth_date = format_date(self.birth_date)
                if self.death_date:
                    death_date = format_date(self.death_date)
                    if death_date and birth_date:
                        if (death_date.month, death_date.day) < (birth_date.month, birth_date.day):
                            return death_date.year - birth_date.year - 1
                        else:
                            return death_date.year - birth_date.year
                    else:
                        return None
                else:
                    today = datetime.now().date()
                    if birth_date:
                        if (today.month, today.day) < (birth_date.month, birth_date.day):
                            return today.year - birth_date.year - 1
                        else:
                            return today.year - birth_date.year
                    else:
                        return None
        except AttributeError:
            return 'INCORRECT DATE FORMAT'

    @classmethod
    def create_person(cls):
        first_name = input("Enter name: ")
        while not first_name or not first_name.isalpha():
            print('Incorrect entry')
            first_name = input("Enter name: ")
        last_name = input("Enter a last name (or press Enter if there is no last name): ")
        middle_name = input("Enter middle name (or press Enter if middle name is missing): ")
        birth_date = input("Enter date of birth (in dd.mm.yyyy format): ")
        while not format_date(birth_date) or not birth_date:
            print("Incorrect date. Please enter the date in the correct format.")
            birth_date = input("Enter date of birth (in dd.mm.yyyy format): ")
        death_date = input("Enter the date of death (or press Enter if there is no date of death): ")
        while death_date and not format_date(death_date):
            print("Incorrect date. Please enter the date in the correct format or click Enter.")
            death_date = input("Enter the date of death (or press Enter if there is no date of death): ")
        gender = input("Enter gender (m/f): ").lower()
        while gender not in ['m', 'f']:
            print("Incorrect gender. Please enter: 'm' or 'f'.")
            gender = input("Enter gender (m/f): ").lower()

        birth_date = format_date(birth_date)
        death_date = format_date(death_date) if death_date else None

        return cls(first_name.title(), birth_date, gender, last_name.title(), middle_name.title(), death_date)

    @classmethod
    def search_person(cls):
        request = input('Type request for search: ')
        found = False
        for item in cls.PEOPLE_BASE:
            if (request.lower() in item.first_name.lower() or
                    request.lower() in item.last_name.lower() or
                    request.lower() in item.middle_name.lower()):
                found = True
                print(f"{item.number}: {item}")
        if not found:
            print("No matching person found.")

    @classmethod
    def save_to_excel(cls, filename):
        try:
            try:
                wb = openpyxl.load_workbook(filename)
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            sheet = wb.active
            titles = ["Item Number", "First Name", "Last Name", "Middle Name",
                      "Birth Date", "Gender", "Death Date", "Age"]

            if not any(cell.value for cell in sheet[1]):
                for col_num, title in enumerate(titles, start=1):
                    sheet.cell(row=1, column=col_num, value=title)

            item_start_row = 1

            for index, item in enumerate(cls.PEOPLE_BASE, start=item_start_row):
                row_data = [
                    index,
                    item.first_name, item.last_name, item.middle_name,
                    item.birth_date, item.gender,
                    item.death_date if item.death_date else None,
                    item.age
                ]
                sheet.append(row_data)

            wb.save(filename)

        except Exception as e:
            return f"Error: {e}"

    @classmethod
    def load_from_excel(cls, filename):
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            uploaded_persons = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7:
                    number, first_name, last_name, middle_name, birth_date, gender, death_date = row[:7]
                    birth_date = format_date(birth_date)
                    death_date = format_date(death_date) if death_date else None
                    person = Person(first_name, birth_date, gender, last_name, middle_name, death_date)
                    uploaded_persons.append(person)
                else:
                    print("Error: Row does not contain sufficient data")
            for item in uploaded_persons:
                print(item)
        except Exception as e:
            print(f"Error: {e}")


def format_date(dates):
    try:
        if isinstance(dates, date):
            return dates
        delimiters = ['/', '.', ',', ' ', '-']
        for item in delimiters:
            if item in dates:
                break
        else:
            return date

        if item:
            day, month, year = map(int, dates.split(item))
            return datetime(year, month, day).date()
        else:
            parts = dates.split()
            if len(parts) == 3:
                day, month, year = map(int, parts)
                return datetime(year, month, day).date()
            else:
                return date
    except Exception as e:
        return f"Error: {e}"

